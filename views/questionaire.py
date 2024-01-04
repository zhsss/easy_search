import json
import zipfile
from collections import defaultdict

import docxtpl
import qrcode
from docx.shared import Mm
from flask import Blueprint, request, make_response
import os

from openpyxl.workbook import Workbook

from dao.departments import d
from dao.member import m
from dao.questionnaire import q
from pathlib import Path

qs_view = Blueprint("questionnaire", __name__)


@qs_view.route('/questionnaires', methods=['GET', 'POST'])
def generate_questionnaire():
    if request.method == "POST":
        online_num = int(request.args.get('online_num'))
        offline_num = int(request.args.get('offline_num'))
        leader_num = int(request.args.get('leader_num'))

        time_key = q.generate_a_qs(online_num, offline_num, leader_num)
        return time_key
    elif request.method == "GET":
        return json.dumps(q.getall())


@qs_view.route('/documents')
def generate_documents():
    time_string = request.args.get('time_string')
    tokens = q.get_tokens(time_string)

    if not os.path.exists("imgs"):
        os.mkdir("imgs")
    if not os.path.exists("output"):
        os.mkdir("output")
    if not os.path.exists(f"imgs/{time_string}"):
        os.mkdir(f"imgs/{time_string}")
    if not os.path.exists(f"output/{time_string}"):
        os.mkdir(f"output/{time_string}")

    types = ["00", "01", "10", "11", "2"]
    string_cor = {
        "00": "无人员_教学教研",
        "01": "无人员_党群行政职能部门",
        "10": "有人员_教学教研",
        "11": "有人员_党群行政职能部门",
        "2": "校领导"
    }

    for typ in types:
        if not os.path.exists(f"output/{time_string}/{string_cor[typ]}"):
            os.mkdir(f"output/{time_string}/{string_cor[typ]}")

    ans = zipfile.ZipFile(f"output/{time_string}.zip", 'w')
    for token in tokens:
        img = qrcode.make(data=f"https://smkaohe.cuc.edu.cn/?time_string={time_string}&token={token}")
        token_qrcode_pic_path = f"imgs/{time_string}/{token}.png"
        with open(token_qrcode_pic_path, 'wb') as f:
            img.save(f)

        progress = json.loads(q.get_progress(time_string, token))
        typ = progress['type']

        if typ == "2":
            tpl = docxtpl.DocxTemplate('templates/校领导.docx')
        elif typ[0] == "1":
            tpl = docxtpl.DocxTemplate('templates/有人员.docx')
        elif typ[0] == "0":
            tpl = docxtpl.DocxTemplate('templates/无人员.docx')
        tpl.render({"img": docxtpl.InlineImage(tpl, token_qrcode_pic_path, Mm(50), Mm(50))})
        tpl.save(f"output/{time_string}/{string_cor[typ]}/{token}.docx")

        ans.write(f"output/{time_string}/{string_cor[typ]}/{token}.docx", f"{string_cor[typ]}/{token}.docx",
                  zipfile.ZIP_DEFLATED)
    ans.close()

    with open(f"output/{time_string}.zip", 'rb') as f:
        data = f.read()
    response = make_response(data)
    response.headers['Content-Disposition'] = 'attachment; filename=ans.zip'

    return response


@qs_view.route('/progress')
def query_progress():
    time_string = request.args.get('time_string')
    token = request.args.get('token')
    progress = q.get_progress(time_string, token)
    return progress


@qs_view.route('/tokens')
def get_all_tokens():
    time_string = request.args.get('time_string')
    return json.dumps(q.get_tokens(time_string))


@qs_view.route('/rate', methods=["POST"])
def rate():
    # 打分
    time_string = request.args.get('time_string')
    token = request.args.get('token')
    did = int(request.args.get('id'))
    score = int(request.args.get('score'))
    no = int(request.args.get('no'))
    q.update_score(time_string, token, did, score, int(no))
    return ""


@qs_view.route('/export', methods=['GET'])
def export():
    def is_progress_done(progress):
        for progress in progress["progress"]:
            if 0 in progress["score"] or 0 in progress["members"].values():
                return False
        return True

    time_string = request.args.get('time_string')
    tokens = q.get_tokens(time_string)

    all_depts = [set(), set()]
    members = defaultdict(set)

    token_process_mp = {
        token: json.loads(q.get_progress(time_string, token)) for token in tokens
    }

    for token, progress in token_process_mp.items():
        typ = progress["type"]
        if typ == "2": continue

        for progress in progress["progress"]:
            all_depts[int(typ[1])].add(progress["name"])
            for member in progress["members"]:
                members[progress["name"]].add(member)

    all_progress = [json.loads(q.get_progress(time_string, token)) for token in tokens]

    target_export_items = {
        "0": "考核组评委",
        "1": "学校评委团",
        "2": "校领导"
    }

    ranked_all_depts = json.loads(d.getall())
    ranked_depts = ranked_all_depts[0] + ranked_all_depts[1]
    dept_rank = defaultdict(lambda: -1)
    for rank, dept in enumerate(ranked_depts):
        if dept not in dept_rank:
            dept_rank[dept] = rank

    '''
       部门成绩 - 3种
       人员打分 - 2种 
    '''

    ans = zipfile.ZipFile(f"output/{time_string}_score.zip", 'w')
    # 生成部门打分
    Path(f"output/{time_string}_score/部门成绩/").mkdir(parents=True, exist_ok=True)
    Path(f"output/{time_string}_score/人员成绩/").mkdir(parents=True, exist_ok=True)
    for target, wb_name in target_export_items.items():
        n_cor = defaultdict(int)
        score_cor = {}
        for dept in all_depts[1]:
            score_cor[dept] = [0] * 2
        for dept in all_depts[0]:
            score_cor[dept] = [0] * 4

        for dept, ms in members.items():
            for member in ms:
                score_cor[f"{dept}-{member}"] = [0]

        dept_wb = Workbook()
        dept_ac = dept_wb.active
        dept_ac.append(["部门", "年度综合评分", "贡献评分", "改革创新评分", "年度进步评分"])
        member_wb = Workbook()
        member_ac = member_wb.active
        member_ac.append(["部门", "人员", "年度综合评分"])

        depts = set()
        valid_progresses = list(filter(lambda p: p["type"][0] == target, all_progress))
        for p in valid_progresses:
            for progress in p["progress"]:
                depts.add(progress["name"])

        all_done_progress_number = 0
        for p in valid_progresses:
            if not is_progress_done(p):
                continue
            all_done_progress_number += 1
            for progress in p["progress"]:
                if 0 not in progress["score"]:
                    for i, score in enumerate(progress["score"]):
                        score_cor[progress["name"]][i] += score
                    n_cor[progress["name"]] += 1
                    for member, score in progress["members"].items():
                        if score != 0:
                            score_cor[f"{progress['name']}-{member}"][0] += score
                            n_cor[f"{progress['name']}-{member}"] += 1
        sorted_depts = sorted(depts, key=lambda dept: dept_rank[dept])

        for dept in sorted_depts:
            dept_ac.append([dept] + [i / (n_cor[dept] or 1) for i in score_cor[dept]])
            for member in members[dept]:
                member_ac.append([dept, member] + [i / (n_cor[f"{dept}-{member}"] or 1) for i in
                                                   score_cor[f"{dept}-{member}"]])

        for i, row in enumerate(dept_ac.rows):
            for j, col in enumerate(row):
                if i == 0 or j == 0: continue
                col.number_format = "0.00_);[Red]\(0.00\)"

        for i, row in enumerate(member_ac.rows):
            for j, col in enumerate(row):
                if i == 0 or j == 0: continue
                col.number_format = "0.00_);[Red]\(0.00\)"

        dept_ac["G1"] = f"打分人数：{all_done_progress_number}"
        member_ac["E1"] = f"打分人数：{all_done_progress_number}"

        dept_wb.save(f"output/{time_string}_score/部门成绩/{wb_name}.xlsx")
        ans.write(f"output/{time_string}_score/部门成绩/{wb_name}.xlsx", f"部门成绩/{wb_name}.xlsx",
                  zipfile.ZIP_DEFLATED)
        if target != "0":
            member_wb.save(f"output/{time_string}_score/人员成绩/{wb_name}.xlsx")
            ans.write(f"output/{time_string}_score/人员成绩/{wb_name}.xlsx", f"人员成绩/{wb_name}.xlsx",
                      zipfile.ZIP_DEFLATED)

    # Path(f"output/{time_string}_score/人员成绩明细/").mkdir(parents=True, exist_ok=True)
    single_member_cor_score = defaultdict(list)
    for p in all_progress:
        for progress in p["progress"]:
            for member, score in progress["members"].items():
                single_member_cor_score[f"{progress['name']}-{member}"].append(score)

    single_wb = Workbook()
    single_ac = single_wb.active
    single_ac.append(["部门", "人员", "年度综合评分"])
    for info, scores in single_member_cor_score.items():
        print(info)
        dept, member = info.split('-')
        single_ac.append([dept, member] + scores)
    for i, row in enumerate(single_ac.rows):
        for j, col in enumerate(row):
            if i == 0 or j == 0: continue
            col.number_format = "0.00_);[Red]\(0.00\)"
    single_wb.save(f"output/{time_string}_score/人员成绩明细.xlsx")
    ans.write(f"output/{time_string}_score/人员成绩明细.xlsx", f"人员成绩明细.xlsx",
              zipfile.ZIP_DEFLATED)

    ans.close()
    with open(f"output/{time_string}_score.zip", 'rb') as f:
        data = f.read()
    response = make_response(data)
    response.headers['Content-Disposition'] = 'attachment; filename=ans.zip'

    return response


@qs_view.route('/questionnaire_status', methods=['GET', 'POST'])
def questionnaire_status():
    if request.method == "POST":
        time_string = request.args.get('time_string')
        token = request.args.get('token')
        status = request.args.get('status')

        if status == '1':
            progress = json.loads(q.get_progress(time_string, token))

            scores = set()
            for p in progress['progress']:
                for s in p['score']:
                    scores.add(s)
                for v in p['members']:
                    scores.add(v)

            if 0 in scores:
                return '0'

        q.set_status(time_string, token, status)
        return "1"
    elif request.method == "GET":
        time_string = request.args.get('time_string')
        token = request.args.get('token')

        return q.get_status(time_string, token) or "0"
