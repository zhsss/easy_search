import json
from collections import defaultdict
import time
from openpyxl import Workbook, load_workbook
from requests import get, post
import sys

if len(sys.argv)<3:
    print('需要命令行参数，第一个为部门xlsx路径，第二个为人员xlsx路径')
    exit(1)

def MaxMatchScore(strA:str,strB:str):
    #！需要把括号去掉匹配以防问题
    strA=strA.split('（')[0]
    strB=strB.split('（')[0]
    matchScore=0
    for c in strA:
        if c in strB:
            matchScore+=1
    return matchScore/( max(len(strA),len(strB)) )

matchCache={}
def MaxMatchBestFit(targets:list[str],patt:str):
    global matchCache
    if patt in matchCache:
        return matchCache[patt]
    
    scoreList=[]
    for tar in targets:
        scoreList.append(MaxMatchScore(tar,patt))
    matchedStr=targets[ scoreList.index( max(scoreList) ) ]
    matchCache[patt]=matchedStr
    return matchedStr
    

#dpart
DEPT_DANG_COL=0
DEPT_ZHENG_COL=1
DEPT_XUE_COL=2
DEPT_SHU_COL=3

DEPT_START_ROW=3
DEPT_END__ROW=23+1

deptXL=load_workbook(sys.argv[1])
deptAc=deptXL.active
deptList=[]
postDeptGroup=[[],[]]

rowIdx=-1
for rowElms in deptAc.rows:
    rowIdx+=1
    if rowIdx<DEPT_START_ROW or rowIdx >=DEPT_END__ROW:
        continue
    colIdx=0
    for dept in rowElms:
        #print(dept)
        dept=dept.value
        if not dept: 
            continue
        dept=dept.strip().split('.')[-1]
        deptList.append(dept)
        if colIdx>DEPT_ZHENG_COL:
            postDeptGroup[0].append(dept)
        else:
            postDeptGroup[1].append(dept)
        colIdx+=1

#member
MEM_DEPT_COL=1
MEM_NAME_COL=2

MEM_START_ROW=2
MEM_END__ROW=240+1

MEM_DEPT_SPLITTER='、'

memData = load_workbook(sys.argv[2])
memAc = memData.active
postDeptMems=defaultdict(list)

rowIdx=-1
lastRowELms=None
for rowElms in memAc.rows:
    rowIdx+=1
    if rowIdx<MEM_START_ROW or rowIdx >=MEM_END__ROW:
        continue
    #print(rowElms[MEM_NAME_COL])
    memDepts=rowElms[MEM_DEPT_COL].value.strip()
    memName=''
    #防止出现多行共用一个名
    if lastRowELms and (not rowElms[MEM_NAME_COL].value):
        memName=lastRowELms[MEM_NAME_COL].value.strip()
    else:
        lastRowELms=rowElms
        memName=rowElms[MEM_NAME_COL].value.strip()
    for memDept in memDepts.split(MEM_DEPT_SPLITTER):
        realDept=MaxMatchBestFit(deptList,memDept)
        postDeptMems[realDept].append(memName)

print(postDeptGroup)
print(postDeptMems)
#exit(0)

# set_depts_api = "http://47.111.143.45:8000/depts"
set_depts_api = "http://202.205.22.179:8000/depts"

post(set_depts_api, data=json.dumps(postDeptGroup, ensure_ascii=False).encode())
# set_member_api = "http://47.111.143.45:8000/members?dept="
set_member_api = "http://202.205.22.179:8000/members?dept="
for dept in deptList:
    post(set_member_api + dept, data=json.dumps(postDeptMems[dept], ensure_ascii=False).encode())
